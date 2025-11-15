# начало файла report_generator.py
import calendar
import os
import re
import tkinter as tk
from openpyxl import load_workbook
from datetime import datetime, timedelta
from tkinter import ttk, messagebox, filedialog
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment


def sanitize_filename(filename):
    """Удаляет недопустимые символы из имени файла."""
    return re.sub(r'[<>:"/\\|?*]', '', filename)

class ReportGeneratorFrame(ttk.LabelFrame):
    def __init__(self, parent, data, cart_weights, detect_anomalies, boiler_app=None):
        super().__init__(parent, text="Формирование отчётов")
        self.parent = parent
        self.data = data
        self.cart_weights = cart_weights
        self.detect_anomalies = detect_anomalies
        self.boiler_app = boiler_app  # Для доступа к рек_лимитам
        self.create_widgets()

    def create_widgets(self):
        # Выбор котельной
        self.boiler_label = ttk.Label(self, text="Котельная:")
        self.boiler_label.grid(row=0, column=0, padx=5, pady=5)
        self.boiler_var = tk.StringVar()
        boiler_options = list(self.cart_weights.keys())
        self.boiler_combobox = ttk.Combobox(self, textvariable=self.boiler_var, values=boiler_options)
        self.boiler_combobox.grid(row=0, column=1, padx=5, pady=5)

        # Выбор периода
        ttk.Label(self, text="Начальная дата:").grid(row=1, column=0, padx=5, pady=5)
        self.start_date_var = tk.StringVar()
        ttk.Entry(self, textvariable=self.start_date_var).grid(row=1, column=1, padx=5, pady=5)
        ttk.Label(self, text="Конечная дата:").grid(row=2, column=0, padx=5, pady=5)
        self.end_date_var = tk.StringVar()
        ttk.Entry(self, textvariable=self.end_date_var).grid(row=2, column=1, padx=5, pady=5)

        # Выбор типа отчёта (не показывать для газовой котельной)
        self.report_type_label = ttk.Label(self, text="Тип отчёта:")
        self.report_type_label.grid(row=3, column=0, padx=5, pady=5)
        self.report_type_var = tk.StringVar()
        report_types = [
            "Полный отчёт",
            "Температуры",
            "Расход угля",
            "Расход воды",
            "Расход на комплексоне",
            "Расход электроэнергии",
            "Сравнительный отчёт по расходу угля",
            "Сравнительный отчёт по электроэнергии",
            "Сравнительный отчёт по общей воде",
         #  "Сравнительный отчёт по воде через комплексон",
            "Сравнительный отчёт по Гкал"
        ]
        self.report_type_combobox = ttk.Combobox(self, textvariable=self.report_type_var, values=report_types, width=35)
        self.report_type_combobox.grid(row=3, column=1, padx=10, pady=5)

        # Привязываем событие изменения котельной
        self.boiler_var.trace_add('write', self.on_boiler_change)

        # Кнопка формирования отчёта
        ttk.Button(self, text="Сформировать отчёт", command=self.generate_report).grid(row=4, column=0, columnspan=2, pady=10)

    def on_boiler_change(self, *args):
        """Обработчик изменения выбора котельной"""
        report_type = self.report_type_var.get()
        if report_type in [
            "Сравнительный отчёт по расходу угля",
            "Сравнительный отчёт по электроэнергии",
            "Сравнительный отчёт по общей воде",
            "Сравнительный отчёт по воде через комплексон",
            "Сравнительный отчёт по Гкал"
        ]:
            # Скрываем выбор котельной и тип отчёта
            self.boiler_label.grid_remove()
            self.boiler_combobox.grid_remove()
            self.report_type_label.grid()
            self.report_type_combobox.grid()
        else:
            # Показываем выбор котельной
            self.boiler_label.grid()
            self.boiler_combobox.grid()
            if self.boiler_var.get() == "Котельная №13":
                self.report_type_label.grid_remove()
                self.report_type_combobox.grid_remove()
            else:
                self.report_type_label.grid()
                self.report_type_combobox.grid()

    def read_gcal_from_json(self, file_path):
        """
        Извлекает расход тепловой энергии (Гкал) из JSON-файла "Потребление ресурса.json".
        Возвращает словарь: {"Котельная №6": 22.77, "Котельная №9": 30.34, ...}
        """
        import json
        import re

        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось открыть JSON-файл: {e}")
            return {}

        # Проверяем структуру файла
        if "Потребление_ресурса" not in data:
            messagebox.showerror("Ошибка", "Неверная структура JSON: отсутствует ключ 'Потребление_ресурса'")
            return {}

        records = data["Потребление_ресурса"]
        results = {}
        current_boiler = None
        inside_target_section = False

        for record in records:
            report_name = record.get("HeaderBand_text_ReportName", "").strip()

            # 1. Определяем котельную
            boiler_match = re.search(r"Котельная\s*№\s*(\d+)", report_name)
            if boiler_match:
                current_boiler = f"Котельная №{boiler_match.group(1)}"
                continue

            # 2. Проверяем точку измерения
            if report_name == "Точка измерения: Тв1; Ресурс: ТС":
                if current_boiler is not None:
                    inside_target_section = True
                continue

            # 3. Если вышли из секции (встретили новую точку или новый объект)
            if ("Точка измерения:" in report_name and report_name != "Точка измерения: Тв1; Ресурс: ТС") or \
                    ("Котельная №" in report_name):
                inside_target_section = False

            # 4. Если внутри нужной секции и нашли "Итого:"
            if inside_target_section and report_name == "Итого:":
                qtv_value = record.get("HeaderBand_text_HeaderColumn_TV_Qtv", "").strip()
                if qtv_value and qtv_value not in ("---", ""):
                    try:
                        value = float(qtv_value.replace(",", "."))
                        results[current_boiler] = value
                    except ValueError:
                        pass  # некорректное значение — пропускаем
                inside_target_section = False  # завершаем обработку секции

        return results

    def generate_comparative_gcal_report(self, start_date, end_date):
        """Генерирует сравнительный отчёт по Гкал на основе JSON-файла"""
        # Открываем диалог выбора файла
        file_path = filedialog.askopenfilename(
            title="Выберите файл 'Потребление ресурса.json'",
            filetypes=[("JSON файлы", "*.json"), ("Все файлы", "*.*")]
        )
        if not file_path:
            return

        # Читаем Гкал из JSON
        gcal_data = self.read_gcal_from_json(file_path)
        if not gcal_data:
            messagebox.showinfo("Информация", "Не найдены данные о Гкал в файле.")
            return

        # Создаём Excel-файл
        wb = Workbook()
        ws = wb.active
        ws.title = "Сравнительный отчёт по Гкал"

        # Стили
        green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        alignment_wrap = Alignment(wrap_text=True)

        # Заголовки
        ws.append(["Сравнительный анализ расхода Гкал"])
        ws.append([f"Период: {start_date.strftime('%d-%m-%Y')} – {end_date.strftime('%d-%m-%Y')}"])
        ws.append(["Теплоисточник", "По РЭК", "По факту", "Экономия"])

        total_rek = 0
        total_fact = 0
        total_economy = 0

        # Все угольные котельные (кроме №13)
        coal_boilers = [b for b in self.cart_weights.keys() if b != "Котельная №13"]

        for boiler in coal_boilers:
            # Получаем фактическое значение из JSON
            fact = gcal_data.get(boiler, 0)

            # Получаем лимит РЭК
            rek = self.calculate_gcal_rek_limit_for_period(boiler, start_date, end_date)

            # Расчёт экономии
            economy = rek - fact
            economy_pct = (economy / rek * 100) if rek != 0 else 0

            ws.append([
                boiler,
                round(rek, 2),
                round(fact, 2),
                f"{round(economy, 3)} ({round(economy_pct, 1)}%)"
            ])
            last_row = ws.max_row
            ws.cell(row=last_row, column=4).alignment = alignment_wrap
            # Подсветка экономии зелёным
            if economy > 0:
                ws.cell(row=last_row, column=4).fill = green_fill

            total_rek += rek
            total_fact += fact
            total_economy += economy

        # Итоговая строка
        total_economy_pct = (total_economy / total_rek * 100) if total_rek != 0 else 0
        ws.append([
            "Итого",
            round(total_rek, 2),
            round(total_fact, 2),
            f"{round(total_economy, 3)} ({round(total_economy_pct, 1)}%)"
        ])
        last_row = ws.max_row
        ws.cell(row=last_row, column=4).alignment = alignment_wrap
        if total_economy > 0:
            ws.cell(row=last_row, column=4).fill = green_fill

        # Автоширина и высота
        self._finalize_worksheet(ws)

        # Сохранение
        today = datetime.now().strftime("%d.%m.%Y")
        file_name = f"Сравнительный отчёт по Гкал {today}.xlsx"
        reports_dir = "reports"
        os.makedirs(reports_dir, exist_ok=True)
        file_path = os.path.join(reports_dir, file_name)
        try:
            wb.save(file_path)
            messagebox.showinfo("Успешно!", f"Отчёт сохранён: {file_path}")
        except PermissionError:
            messagebox.showerror("Ошибка", "Файл открыт в другой программе.")

    def _finalize_worksheet(self, ws):
        """Автоширина и высота строк"""
        # Ширина
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    max_len = max(max_len, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max_len + 2

        # Высота
        for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
            max_height = 0
            for cell in row:
                if isinstance(cell.value, str) and "\n" in cell.value:
                    line_count = cell.value.count("\n") + 1
                    max_height = max(max_height, line_count * 12)
            if max_height > 0:
                ws.row_dimensions[row[0].row].height = max_height

    def _save_comparative_report(self, wb, base_name, start_date, end_date):
        file_name = f"{base_name} {start_date.strftime('%d-%m-%Y')} - {end_date.strftime('%d-%m-%Y')}.xlsx"
        reports_dir = "reports"
        os.makedirs(reports_dir, exist_ok=True)
        file_path = os.path.join(reports_dir, file_name)
        try:
            wb.save(file_path)
            messagebox.showinfo("Успешно!", f"Отчёт сохранён: {file_path}")
        except PermissionError:
            messagebox.showerror("Ошибка", "Файл открыт в другой программе.")

    def generate_comparative_electricity_report(self, start_date, end_date):
        wb = Workbook()
        ws = wb.active
        ws.title = "Сравнительный отчёт по эл."
        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        alignment_wrap = Alignment(wrap_text=True)

        # Заголовки
        ws.append(["Сравнительный отчёт по расходу электроэнергии"])
        ws.append([f"Период: {start_date.strftime('%d-%m-%Y')} – {end_date.strftime('%d-%m-%Y')}"])
        ws.append(["Теплоисточник", "По РЭК", "По факту", "Перерасход", "Экономия"])

        total_rek = 0
        total_fact = 0
        total_over = 0
        total_economy = 0

        coal_boilers = [b for b in self.cart_weights.keys() if b != "Котельная №13"]
        for boiler in coal_boilers:
            boiler_data = [
                row for row in self.data
                if row["boiler"] == boiler and start_date <= datetime.strptime(row["date"], "%d-%m-%Y") <= end_date
            ]
            if not boiler_data:
                continue

            # Расход электроэнергии = разница между последним и первым показанием
            sorted_data = sorted(boiler_data, key=lambda x: datetime.strptime(x["date"], "%d-%m-%Y"))
            first_reading = sorted_data[0]["electricity"]
            last_reading = sorted_data[-1]["electricity"]
            fact = last_reading - first_reading

            rek = self.calculate_electr_rek_limit_for_period(boiler, start_date, end_date)
            over = max(0, fact - rek)
            economy = max(0, rek - fact)

            over_pct = (over / rek * 100) if rek != 0 else 0
            economy_pct = (economy / rek * 100) if rek != 0 else 0

            ws.append([
                boiler,
                round(rek, 3),
                round(fact, 3),
                f"{round(over, 3)} ({round(over_pct, 1)}%)",
                f"{round(economy, 3)} ({round(economy_pct, 1)}%)"
            ])
            last_row = ws.max_row
            ws.cell(row=last_row, column=4).alignment = alignment_wrap
            ws.cell(row=last_row, column=5).alignment = alignment_wrap

            total_rek += rek
            total_fact += fact
            total_over += over
            total_economy += economy

        # Итого
        total_over_pct = (total_over / total_rek * 100) if total_rek != 0 else 0
        total_economy_pct = (total_economy / total_rek * 100) if total_rek != 0 else 0
        ws.append([
            "Итого",
            round(total_rek, 3),
            round(total_fact, 3),
            f"{round(total_over, 3)} ({round(total_over_pct, 1)}%)",
            f"{round(total_economy, 3)} ({round(total_economy_pct, 1)}%)"
        ])
        last_row = ws.max_row
        ws.cell(row=last_row, column=4).alignment = alignment_wrap
        ws.cell(row=last_row, column=5).alignment = alignment_wrap

        # Подсветка при перерасходе
        for row_idx in range(4, ws.max_row):  # строки с данными (начиная с 4-й)
            if ws.cell(row=row_idx, column=3).value > ws.cell(row=row_idx, column=2).value:
                for col in range(2, 6):
                    ws.cell(row=row_idx, column=col).fill = red_fill

        self._finalize_worksheet(ws)
        self._save_comparative_report(wb, "Сравнительный отчёт по электроэнергии", start_date, end_date)

    def generate_comparative_feed_water_report(self, start_date, end_date):
        wb = Workbook()
        ws = wb.active
        ws.title = "Сравнительный отчёт по компл."
        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        alignment_wrap = Alignment(wrap_text=True)

        ws.append(["Сравнительный отчёт по расходу воды через комплексон"])
        ws.append([f"Период: {start_date.strftime('%d-%m-%Y')} – {end_date.strftime('%d-%m-%Y')}"])
        ws.append(["Теплоисточник", "По РЭК", "По факту", "Перерасход", "Экономия"])

        total_rek = 0
        total_fact = 0
        total_over = 0
        total_economy = 0

        coal_boilers = [b for b in self.cart_weights.keys() if b != "Котельная №13"]
        for boiler in coal_boilers:
            boiler_data = [
                row for row in self.data
                if row["boiler"] == boiler and start_date <= datetime.strptime(row["date"], "%d-%m-%Y") <= end_date
            ]
            if not boiler_data:
                continue

            # feed_water — это уже расход за день → суммируем
            fact = sum(row["feed_water"] for row in boiler_data)
            rek = self.calculate_feed_water_rek_limit_for_period(boiler, start_date, end_date)
            over = max(0, fact - rek)
            economy = max(0, rek - fact)

            over_pct = (over / rek * 100) if rek != 0 else 0
            economy_pct = (economy / rek * 100) if rek != 0 else 0

            ws.append([
                boiler,
                round(rek, 3),
                round(fact, 3),
                f"{round(over, 3)} ({round(over_pct, 1)}%)",
                f"{round(economy, 3)} ({round(economy_pct, 1)}%)"
            ])
            last_row = ws.max_row
            ws.cell(row=last_row, column=4).alignment = alignment_wrap
            ws.cell(row=last_row, column=5).alignment = alignment_wrap

            total_rek += rek
            total_fact += fact
            total_over += over
            total_economy += economy

        # Итого
        total_over_pct = (total_over / total_rek * 100) if total_rek != 0 else 0
        total_economy_pct = (total_economy / total_rek * 100) if total_rek != 0 else 0
        ws.append([
            "Итого",
            round(total_rek, 3),
            round(total_fact, 3),
            f"{round(total_over, 3)} ({round(total_over_pct, 1)}%)",
            f"{round(total_economy, 3)} ({round(total_economy_pct, 1)}%)"
        ])
        last_row = ws.max_row
        ws.cell(row=last_row, column=4).alignment = alignment_wrap
        ws.cell(row=last_row, column=5).alignment = alignment_wrap

        # Подсветка перерасхода
        for row_idx in range(4, ws.max_row):
            if ws.cell(row=row_idx, column=3).value > ws.cell(row=row_idx, column=2).value:
                for col in range(2, 6):
                    ws.cell(row=row_idx, column=col).fill = red_fill

        self._finalize_worksheet(ws)
        self._save_comparative_report(wb, "Сравнительный отчёт по воде через комплексон", start_date, end_date)


    def generate_comparative_water_report(self, start_date, end_date):
        wb = Workbook()
        ws = wb.active
        ws.title = "Сравнительный отчёт по воде"
        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        alignment_wrap = Alignment(wrap_text=True)

        ws.append(["Сравнительный отчёт по расходу воды"])
        ws.append([f"Период: {start_date.strftime('%d-%m-%Y')} – {end_date.strftime('%d-%m-%Y')}"])
        ws.append(["Теплоисточник", "По РЭК", "По факту", "Перерасход", "Экономия"])

        total_rek = 0
        total_fact = 0
        total_over = 0
        total_economy = 0

        coal_boilers = [b for b in self.cart_weights.keys() if b != "Котельная №13"]
        for boiler in coal_boilers:
            boiler_data = [
                row for row in self.data
                if row["boiler"] == boiler and start_date <= datetime.strptime(row["date"], "%d-%m-%Y") <= end_date
            ]
            if not boiler_data:
                continue

            sorted_data = sorted(boiler_data, key=lambda x: datetime.strptime(x["date"], "%d-%m-%Y"))
            first_reading = sorted_data[0]["total_water"]
            last_reading = sorted_data[-1]["total_water"]
            fact = last_reading - first_reading

            rek = self.calculate_water_rek_limit_for_period(boiler, start_date, end_date)
            over = max(0, fact - rek)
            economy = max(0, rek - fact)

            over_pct = (over / rek * 100) if rek != 0 else 0
            economy_pct = (economy / rek * 100) if rek != 0 else 0

            ws.append([
                boiler,
                round(rek, 3),
                round(fact, 3),
                f"{round(over, 3)} ({round(over_pct, 1)}%)",
                f"{round(economy, 3)} ({round(economy_pct, 1)}%)"
            ])
            last_row = ws.max_row
            ws.cell(row=last_row, column=4).alignment = alignment_wrap
            ws.cell(row=last_row, column=5).alignment = alignment_wrap

            total_rek += rek
            total_fact += fact
            total_over += over
            total_economy += economy

        # Итого
        total_over_pct = (total_over / total_rek * 100) if total_rek != 0 else 0
        total_economy_pct = (total_economy / total_rek * 100) if total_rek != 0 else 0
        ws.append([
            "Итого",
            round(total_rek, 3),
            round(total_fact, 3),
            f"{round(total_over, 3)} ({round(total_over_pct, 1)}%)",
            f"{round(total_economy, 3)} ({round(total_economy_pct, 1)}%)"
        ])
        last_row = ws.max_row
        ws.cell(row=last_row, column=4).alignment = alignment_wrap
        ws.cell(row=last_row, column=5).alignment = alignment_wrap

        # Подсветка перерасхода
        for row_idx in range(4, ws.max_row):
            if ws.cell(row=row_idx, column=3).value > ws.cell(row=row_idx, column=2).value:
                for col in range(2, 6):
                    ws.cell(row=row_idx, column=col).fill = red_fill

        self._finalize_worksheet(ws)
        self._save_comparative_report(wb, "Сравнительный отчёт по воде", start_date, end_date)


    def generate_comparative_coal_report(self, start_date, end_date):
        wb = Workbook()
        ws = wb.active
        ws.title = "Сравнительный отчёт по углю"
        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

        ws.append(["Сравнительный отчёт по расходу угля"])
        ws.append(["Теплоисточник", "По РЭК", "По факту", "Шлак", "Экономия"])

        total_rek = 0
        total_fact = 0
        total_slag = 0
        total_economy = 0

        # Все угольные котельные (кроме №13)
        coal_boilers = [b for b in self.cart_weights.keys() if b != "Котельная №13"]

        for boiler in coal_boilers:
            # Фильтруем все данные по этой котельной и периоду
            boiler_data = [
                row for row in self.data
                if row["boiler"] == boiler
                   and start_date <= datetime.strptime(row["date"], "%d-%m-%Y") <= end_date
            ]
            if not boiler_data:
                continue

            # Фактический расход — сумма значений "coal"
            fact = sum(row["coal"] for row in boiler_data)
            rek = self.calculate_coal_rek_limit_for_period(boiler, start_date, end_date)

            # Шлак — это зола (ash), суммируем за период
            slag = sum(row["ash"] for row in boiler_data)
            # Экономия — это разница между лимитом и фактом угля
            economy = max(0, rek - fact)

            slag_pct = (slag / rek * 100) if rek != 0 else 0
            economy_pct = (economy / rek * 100) if rek != 0 else 0

            ws.append([
                boiler,
                round(rek, 3),
                round(fact, 3),
                f"{round(slag, 3)} ({round(slag_pct, 1)}%)",
                f"{round(economy, 3)} ({round(economy_pct, 1)}%)"
            ])

            # Устанавливаем перенос текста для столбцов "Шлак" и "Экономия"
            last_row = ws.max_row
            ws.cell(row=last_row, column=4).alignment = Alignment(wrap_text=True)
            ws.cell(row=last_row, column=5).alignment = Alignment(wrap_text=True)

            total_rek += rek
            total_fact += fact
            total_slag += slag
            total_economy += economy

        # Итоговая строка
        total_slag_pct = (total_slag / total_rek * 100) if total_rek != 0 else 0
        total_economy_pct = (total_economy / total_rek * 100) if total_rek != 0 else 0
        ws.append([
            "Итого",
            round(total_rek, 3),
            round(total_fact, 3),
            f"{round(total_slag, 3)} ({round(total_slag_pct, 1)}%)",
            f"{round(total_economy, 3)} ({round(total_economy_pct, 1)}%)"
        ])


        # ПОДСВЕТКА ПРИ ПЕРЕРАСХОДЕ УГЛЯ
        for row_idx in range(3, ws.max_row):  # Не включая "Итого"
            fact_cell = ws.cell(row=row_idx, column=3)  # По факту
            rek_cell = ws.cell(row=row_idx, column=2)  # По РЭК
            if fact_cell.value > rek_cell.value:
                for col in range(2, 6):  # Подсвечиваем всю строку (столбцы B–E)
                    ws.cell(row=row_idx, column=col).fill = red_fill

        # Автоширина столбцов
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    max_len = max(max_len, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max_len + 2

        for row in ws.iter_rows(min_row=3, max_row=ws.max_row):  # Начинаем с 3-й строки (после заголовков)
            max_height = 0
            for cell in row:
                if isinstance(cell.value, str) and "\n" in cell.value:
                    # Если есть перенос строки, увеличиваем высоту
                    line_count = cell.value.count("\n") + 1
                    max_height = max(max_height, line_count * 10)
            if max_height > 0:
                ws.row_dimensions[row[0].row].height = max_height

        # Сохранение
        file_name = f"Сравнительный отчёт по расходу угля {start_date.strftime('%d-%m-%Y')} - {end_date.strftime('%d-%m-%Y')}.xlsx"
        reports_dir = "reports"
        os.makedirs(reports_dir, exist_ok=True)
        file_path = os.path.join(reports_dir, file_name)
        try:
            wb.save(file_path)
            messagebox.showinfo("Успешно!", f"Отчёт сохранён: {file_path}")
        except PermissionError:
            messagebox.showerror("Ошибка", "Файл открыт в другой программе.")

    def generate_report(self):
        boiler = self.boiler_var.get()
        start_date_str = self.start_date_var.get()
        end_date_str = self.end_date_var.get()
        report_type = self.report_type_var.get()

        if report_type in [
            "Сравнительный отчёт по расходу угля",
            "Сравнительный отчёт по электроэнергии",
            "Сравнительный отчёт по общей воде",
            "Сравнительный отчёт по воде через комплексон",
            "Сравнительный отчёт по Гкал"
        ]:
            try:
                start_date = datetime.strptime(start_date_str, "%d-%m-%Y")
                end_date = datetime.strptime(end_date_str, "%d-%m-%Y")
            except ValueError:
                messagebox.showerror("Ошибка", "Неверный формат даты. Используйте ДД-ММ-ГГГГ.")
                return
            if report_type == "Сравнительный отчёт по расходу угля":
                self.generate_comparative_coal_report(start_date, end_date)
            elif report_type == "Сравнительный отчёт по электроэнергии":
                self.generate_comparative_electricity_report(start_date, end_date)
            elif report_type == "Сравнительный отчёт по общей воде":
                self.generate_comparative_water_report(start_date, end_date)
            elif report_type == "Сравнительный отчёт по воде через комплексон":
                self.generate_comparative_feed_water_report(start_date, end_date)
            elif report_type == "Сравнительный отчёт по Гкал":
                self.generate_comparative_gcal_report(start_date, end_date)
            return

        if report_type == "Сравнительный отчёт по расходу угля":
            # Формируем отчёт по всем угольным котельным
            # Нужно получить start_date и end_date из строки
            try:
                start_date = datetime.strptime(start_date_str, "%d-%m-%Y")
                end_date = datetime.strptime(end_date_str, "%d-%m-%Y")
            except ValueError:
                messagebox.showerror("Ошибка", "Неверный формат даты. Используйте ДД-ММ-ГГГГ.")
                return
            self.generate_comparative_coal_report(start_date, end_date)
            return

        if not all([boiler, start_date_str, end_date_str]):
            messagebox.showerror("Ошибка", "Пожалуйста, заполните все поля.")
            return

        try:
            start_date = datetime.strptime(start_date_str, "%d-%m-%Y")
            end_date = datetime.strptime(end_date_str, "%d-%m-%Y")
        except ValueError:
            messagebox.showerror("Ошибка", "Неверный формат даты. Используйте ДД-ММ-ГГГГ.")
            return

        # Фильтрация данных
        filtered_data = [
            row for row in self.data
            if (
                    row["boiler"] == boiler and
                    start_date <= datetime.strptime(row["date"], "%d-%m-%Y") <= end_date
            )
        ]

        if not filtered_data:
            messagebox.showinfo("Информация", f"Нет данных для {boiler} за указанный период.")
            return

        # Обнаружение аномалий
        anomalies = self.detect_anomalies(filtered_data)

        # Создание Excel-файла
        wb = Workbook()
        ws = wb.active
        ws.title = "Отчёт"

        # Стили для ячеек
        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

        if boiler == "Котельная №13":
            # Таблица 1: показания счётчиков
            ws.append(["Котельная №13"])
            ws.append(["Дата", "Показания счётчика (тыс. м³)"])

            meter_readings = []
            for row in filtered_data:
                gas_value = row.get("gas", 0)
                meter_readings.append((row["date"], gas_value))
                ws.append([row["date"], gas_value])

            # Таблица 2: расчёт расхода и сравнение с лимитом
            ws.append([])
            ws.append(["Расчёт расхода газа"])
            ws.append(["Период",
                       # "Начальный день",
                       # "Посл.день",
                       "Расход газа по факту", "По РЭК", "Экономия"])

            if len(meter_readings) >= 2:
                start_reading = meter_readings[0][1]
                end_reading = meter_readings[-1][1]
                gas_consumption = end_reading - start_reading
                gas_limit = self.calculate_gas_limit_for_period(start_date_str, end_date_str)
                economy = gas_limit - gas_consumption
                economy_percent = (economy / gas_limit * 100) if gas_limit != 0 else 0

                ws.append([
                    f"{meter_readings[0][0]} - {meter_readings[-1][0]}",
                    gas_consumption,
                    gas_limit,
                    f"{round(economy, 3)} | ({round(economy_percent, 1)}%)"
                ])

                # last_row = ws.max_row
                # ws.cell(row=last_row, column=4).alignment = Alignment(wrap_text=True)

                # Подсветка перерасхода
                if economy < 0:
                    ws.cell(row=ws.max_row, column=2).fill = red_fill
                    ws.cell(row=ws.max_row, column=4).fill = red_fill
        else:
            # Для угольных котельных
            if report_type in ["Расход угля", "Расход электроэнергии", "Расход воды", "Расход на комплексоне"]:
                # Таблица 1: Показания
                if report_type == "Расход угля":
                    ws.append(["Показания по расходу угля"])
                    ws.append(["Дата", "Расход угля (кг)"])
                    for row in filtered_data:
                        ws.append([
                            row["date"],
                            row["coal"],
                        ])

                elif report_type == "Расход на комплексоне":
                    ws.append([f"Расход воды через комплексон — {boiler}"])
                    ws.append(["Дата", "Расход (м³)"])
                    for row in filtered_data:
                        ws.append([
                            row["date"],
                            row["feed_water"]
                        ])

                elif report_type == "Расход электроэнергии":
                    ws.append([f"Расход электроэнергии — {boiler}"])
                    ws.append(["Дата", "Показания счётчика (кВт·ч)"])
                    for row in filtered_data:
                        ws.append([
                            row["date"],
                            row["electricity"]
                        ])
                else:  # Расход воды
                    ws.append([f"Показания счётчика воды — {boiler}"])
                    ws.append(["Дата", "Показания счётчика (м³)"])
                    for row in filtered_data:
                        ws.append([
                            row["date"],
                            row["total_water"]
                        ])

                # Таблица 2: Расчёт расхода и сравнение с лимитом
                ws.append([])  # Пустая строка между таблицами
                ws.append(["Расчёт расхода"])

                if report_type == "Расход угля":
                    total_coal = sum(row["coal"] for row in filtered_data)
                    coal_limit = self.calculate_coal_rek_limit_for_period(boiler, start_date, end_date)
                    economy_coal = coal_limit - total_coal
                    economy_percent = (economy_coal / coal_limit * 100) if coal_limit != 0 else 0
                    ws.append(["Период", "Расход угля (кг)", "Лимит РЭК", "Экономия"])
                    ws.append([
                        f"{filtered_data[0]['date']} - {filtered_data[-1]['date']}",
                        total_coal,
                        coal_limit,
                        f"{round(economy_coal, 3)} ({round(economy_percent, 1)}%)"
                    ])
                    if economy_coal < 0:
                        for col in range(2, 5):
                            ws.cell(row=ws.max_row, column=col).fill = red_fill

                elif report_type == "Расход на комплексоне":
                    total_feed_water = sum(row["feed_water"] for row in filtered_data)
                    feed_water_limit = self.calculate_feed_water_rek_limit_for_period(boiler, start_date, end_date)
                    economy_feed = feed_water_limit - total_feed_water
                    economy_percent = (economy_feed / feed_water_limit * 100) if feed_water_limit != 0 else 0
                    ws.append(["Период", "Расход на комплексоне (м³)", "По РЭК", "Экономия"])
                    ws.append([
                        f"{filtered_data[0]['date']} - {filtered_data[-1]['date']}",
                        total_feed_water,
                        feed_water_limit,
                        f"{round(economy_feed, 3)} ({round(economy_percent, 1)}%)"
                    ])
                    if economy_feed < 0:
                        for col in range(2, 5):
                            ws.cell(row=ws.max_row, column=col).fill = red_fill

                elif report_type == "Расход электроэнергии":
                    sorted_data = sorted(filtered_data, key=lambda x: datetime.strptime(x["date"], "%d-%m-%Y"))
                    first_reading = sorted_data[0]["electricity"]
                    last_reading = sorted_data[-1]["electricity"]
                    electricity_consumption = last_reading - first_reading
                    electricity_limit = self.calculate_electr_rek_limit_for_period(boiler, start_date, end_date)
                    economy_electricity = electricity_limit - electricity_consumption
                    economy_percent = (economy_electricity / electricity_limit * 100) if electricity_limit != 0 else 0
                    ws.append(["Период", "Расход электроэнергии", "По РЭК", "Экономия"])
                    ws.append([
                        f"{sorted_data[0]['date']} - {sorted_data[-1]['date']}",
                        electricity_consumption,
                        electricity_limit,
                        f"{round(economy_electricity, 3)} ({round(economy_percent, 1)}%)"
                    ])
                    if economy_electricity < 0:
                        for col in range(2, 5):
                            ws.cell(row=ws.max_row, column=col).fill = red_fill

                else:  # Расход воды
                    sorted_data = sorted(filtered_data, key=lambda x: datetime.strptime(x["date"], "%d-%m-%Y"))
                    first_reading = sorted_data[0]["total_water"]
                    last_reading = sorted_data[-1]["total_water"]
                    water_consumption = last_reading - first_reading
                    water_limit = self.calculate_water_rek_limit_for_period(boiler, start_date, end_date)
                    economy_water = water_limit - water_consumption
                    economy_percent = (economy_water / water_limit * 100) if water_limit != 0 else 0
                    ws.append(["Период", "Расход воды (м³)", "По РЭК", "Экономия"])
                    ws.append([
                        f"{sorted_data[0]['date']} - {sorted_data[-1]['date']}",
                        water_consumption,
                        water_limit,
                        f"{round(economy_water, 3)} ({round(economy_percent, 1)}%)"
                    ])
                    if economy_water < 0:
                        for col in range(2, 5):
                            ws.cell(row=ws.max_row, column=col).fill = red_fill

                # Подсветка аномалий в первой таблице
                for row_idx, row in enumerate(filtered_data, start=3):
                    for anomaly in anomalies:
                        for anomaly_row_idx, anomaly_keys in anomaly.items():
                            if row_idx == anomaly_row_idx + 3:
                                for key in anomaly_keys:
                                    if report_type == "Расход угля" and key in ["coal", "ash"]:
                                        col_idx = ["coal", "ash"].index(key) + 2
                                        ws.cell(row=row_idx, column=col_idx).fill = red_fill
                                    elif report_type == "Расход электроэнергии" and key == "electricity":
                                        ws.cell(row=row_idx, column=2).fill = red_fill
                                    elif report_type == "Расход воды" and key == "total_water":
                                        ws.cell(row=row_idx, column=2).fill = red_fill
                                    elif report_type == "Расход на комплексоне" and key == "feed_water":
                                        ws.cell(row=row_idx, column=2).fill = red_fill
            else:
                # Остальные типы отчётов (Температуры, Расход воды и т.д.) остаются без изменений
                if report_type == "Полный отчёт":
                    headers = ["Дата",
                               # "Котельная",
                               "Расход электроэнергии", "Общий расход воды",
                               "Расход на комплексоне", "Расход угля (кг)", "Зола (кг)",
                               "Температура подачи воды", "Температура обратной подачи воды",
                               "Температура наружного воздуха"]
                elif report_type == "Температуры":
                    headers = ["Дата",
                               # "Котельная",
                               "Температура подачи воды",
                               "Температура обратной подачи воды", "Температура наружного воздуха"]
                elif report_type == "Расход на комплексоне":
                    headers = ["Дата",
                               # "Котельная",
                               "Расход на комплексоне"]

                ws.append(headers)

                for row_idx, row in enumerate(filtered_data, start=2):
                    if report_type == "Полный отчёт":
                        data_row = [
                            row["date"],
                            # row["boiler"],
                            row["electricity"],
                            row["total_water"],
                            row["feed_water"],
                            row["coal"],
                            row["ash"],
                            row["supply_temp"],
                            row["return_temp"],
                            row["outdoor_temp"]
                        ]
                    elif report_type == "Температуры":
                        data_row = [
                            row["date"],
                            # row["boiler"],
                            row["supply_temp"],
                            row["return_temp"],
                            row["outdoor_temp"]
                        ]
                    elif report_type == "Расход воды":
                        data_row = [
                            row["date"],
                            # row["boiler"],
                            row["total_water"]
                        ]
                    elif report_type == "Расход на комплексоне":
                        data_row = [
                            row["date"],
                            # row["boiler"],
                            row["feed_water"]
                        ]
                    ws.append(data_row)


                    # Подсветка аномалий
                    for anomaly in anomalies:
                        for anomaly_row_idx, anomaly_keys in anomaly.items():
                            if row_idx == anomaly_row_idx + 2:
                                for key in anomaly_keys:
                                    header = {
                                        "date": "Дата",
                                        # "boiler": "Котельная",
                                        "electricity": "Расход электроэнергии",
                                        "total_water": "Общий расход воды",
                                        "feed_water": "Расход на комплексоне",
                                        "coal": "Расход угля (кг)",
                                        "ash": "Зола (кг)",
                                        "supply_temp": "Температура подачи воды",
                                        "return_temp": "Температура обратной подачи воды",
                                        "outdoor_temp": "Температура наружного воздуха"
                                    }.get(key)
                                    if header and header in headers:
                                        col_idx = headers.index(header) + 1
                                        ws.cell(row=row_idx, column=col_idx).fill = red_fill

        # Автоматическое выравнивание ширины столбцов
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 1)
            ws.column_dimensions[column].width = adjusted_width

        # Сохранение файла
        sanitized_boiler_name = sanitize_filename(boiler)
        sanitized_report_type = sanitize_filename(report_type)
        file_name = f"{sanitized_boiler_name} {sanitized_report_type} {start_date_str} - {end_date_str}.xlsx"
        reports_dir = "reports"
        os.makedirs(reports_dir, exist_ok=True)
        file_path = os.path.join(reports_dir, file_name)
        try:
            wb.save(file_path)
            messagebox.showinfo("Успешно!", f"Отчёт сохранён в файл: {file_path}")
        except PermissionError:
            messagebox.showerror("Ошибка", f"Не удалось сохранить файл. Возможно, он открыт в другой программе.")

    """РЭК – региональная энергетическая комиссия. 
            Лимиты по РЭК определяют, 
            сколько котельной разрешено потратить того или иного ресурса"""


    def calculate_feed_water_rek_limit_for_period(self, boiler, start_date, end_date):
        """Рассчитывает лимит РЭК на воду через комплексон за указанный период"""
        boiler_data = self.boiler_app.rek_limits["boilers"].get(boiler.replace("Котельная №", ""), {})
        if not boiler_data:
            return 0
        first_half = boiler_data.get("first_half", {}).get("feed_water", 0)
        second_half = boiler_data.get("second_half", {}).get("feed_water", 0)
        month_coefficients = self.boiler_app.rek_limits.get("month_coefficients", {})
        total_limit = 0
        current_date = start_date
        while current_date <= end_date:
            month = current_date.month
            year = current_date.year
            if month in [6, 7, 8]:
                current_date += timedelta(days=1)
                continue
            coeff = month_coefficients.get(str(month), 0)
            if month <= 5:
                half_limit = first_half
            else:
                half_limit = second_half
            _, days_in_month = calendar.monthrange(year, month)
            daily_limit = (half_limit * coeff) / days_in_month
            total_limit += daily_limit
            current_date += timedelta(days=1)
        return round(total_limit, 3)


    def calculate_coal_rek_limit_for_period(self, boiler, start_date, end_date):
        """Рассчитывает лимит РЭК на уголь за указанный период"""
        # Получаем данные из rek_limits.json
        boiler_data = self.boiler_app.rek_limits["boilers"].get(boiler.replace("Котельная №", ""), {})
        if not boiler_data:
            return 0

        first_half = boiler_data.get("first_half", {}).get("coal", 0)
        second_half = boiler_data.get("second_half", {}).get("coal", 0)

        month_coefficients = self.boiler_app.rek_limits.get("month_coefficients", {})

        total_limit = 0
        current_date = start_date

        while current_date <= end_date:
            month = current_date.month
            year = current_date.year

            # Пропускаем летние месяцы (июнь-август)
            if month in [6, 7, 8]:
                current_date += timedelta(days=1)
                continue

            # Получаем коэффициент для текущего месяца
            coeff = month_coefficients.get(str(month), 0)

            # Определяем лимит для полугодия
            if month <= 5:  # Первое полугодие (январь-май)
                half_limit = first_half
            else:  # Второе полугодие (сентябрь-декабрь)
                half_limit = second_half

            # Рассчитываем дневной лимит
            _, days_in_month = calendar.monthrange(year, month)
            daily_limit = (half_limit * coeff) / days_in_month
            total_limit += daily_limit

            current_date += timedelta(days=1)

        return round(total_limit, 3)

    def calculate_electr_rek_limit_for_period(self, boiler, start_date, end_date):
        """Рассчитывает лимит РЭК на электроэнергию за указанный период"""
        # Получаем данные из rek_limits.json
        boiler_data = self.boiler_app.rek_limits["boilers"].get(boiler.replace("Котельная №", ""), {})
        if not boiler_data:
            return 0

        first_half = boiler_data.get("first_half", {}).get("electricity", 0)
        second_half = boiler_data.get("second_half", {}).get("electricity", 0)

        month_coefficients = self.boiler_app.rek_limits.get("month_coefficients", {})

        total_limit = 0
        current_date = start_date

        while current_date <= end_date:
            month = current_date.month
            year = current_date.year

            # Пропускаем летние месяцы (июнь-август)
            if month in [6, 7, 8]:
                current_date += timedelta(days=1)
                continue

            # Получаем коэффициент для текущего месяца
            coeff = month_coefficients.get(str(month), 0)

            # Определяем лимит для полугодия
            if month <= 5:  # Первое полугодие (январь-май)
                half_limit = first_half
            else:  # Второе полугодие (сентябрь-декабрь)
                half_limit = second_half

            # Рассчитываем дневной лимит
            _, days_in_month = calendar.monthrange(year, month)
            daily_limit = (half_limit * coeff) / days_in_month
            total_limit += daily_limit

            current_date += timedelta(days=1)

        return round(total_limit, 3)


    def calculate_gas_limit_for_period(self, start_date_str, end_date_str):
        """рассчёт лимит РЭК на газ за указанный период"""
        try:
            start_date = datetime.strptime(start_date_str, "%d-%m-%Y")
            end_date = datetime.strptime(end_date_str, "%d-%m-%Y")
        except ValueError:
            return 0

        total_days_in_period = (end_date - start_date).days + 1
        gas_year_limit = self.boiler_app.rek_limits["gas_first_half"] + self.boiler_app.rek_limits["gas_second_half"]
        month_coefficients = self.boiler_app.rek_limits.get("month_coefficients", {})

        total_limit = 0.0
        current_date = start_date

        while current_date <= end_date:
            year = current_date.year
            month = current_date.month
            # Получаем количество дней в месяце
            _, month_days = calendar.monthrange(year, month)
            # Получаем коэффициент месяца
            coeff = month_coefficients.get(str(month), 0)
            # Вычисляем лимит на день
            daily_limit = gas_year_limit * coeff / month_days
            # Прибавляем дневной лимит
            total_limit += daily_limit
            current_date += timedelta(days=1)

        return round(total_limit, 3)

    def calculate_water_rek_limit_for_period(self, boiler, start_date, end_date):
        """Рассчитывает лимит РЭК на воду за указанный период"""
        boiler_data = self.boiler_app.rek_limits["boilers"].get(boiler.replace("Котельная №", ""), {})
        if not boiler_data:
            return 0

        first_half = boiler_data.get("first_half", {}).get("water", 0)
        second_half = boiler_data.get("second_half", {}).get("water", 0)

        month_coefficients = self.boiler_app.rek_limits.get("month_coefficients", {})

        total_limit = 0
        current_date = start_date

        while current_date <= end_date:
            month = current_date.month
            year = current_date.year

            if month in [6, 7, 8]:
                current_date += timedelta(days=1)
                continue

            coeff = month_coefficients.get(str(month), 0)

            if month <= 5:
                half_limit = first_half
            else:
                half_limit = second_half

            _, days_in_month = calendar.monthrange(year, month)
            daily_limit = (half_limit * coeff) / days_in_month
            total_limit += daily_limit

            current_date += timedelta(days=1)

        return round(total_limit, 3)

    def calculate_gcal_rek_limit_for_period(self, boiler, start_date, end_date):
        """Рассчитывает лимит РЭК на Гкал (гикакаллории) за указанный период"""
        boiler_data = self.boiler_app.rek_limits["boilers"].get(boiler.replace("Котельная №", ""), {})
        if not boiler_data:
            return 0

        first_half = boiler_data.get("first_half", {}).get("gcal", 0)
        second_half = boiler_data.get("second_half", {}).get("gcal", 0)

        month_coefficients = self.boiler_app.rek_limits.get("month_coefficients", {})

        total_limit = 0
        current_date = start_date

        while current_date <= end_date:
            month = current_date.month
            year = current_date.year

            if month in [6, 7, 8]:
                current_date += timedelta(days=1)
                continue

            coeff = month_coefficients.get(str(month), 0)

            if month <= 5:
                half_limit = first_half
            else:
                half_limit = second_half

            _, days_in_month = calendar.monthrange(year, month)
            daily_limit = (half_limit * coeff) / days_in_month
            total_limit += daily_limit

            current_date += timedelta(days=1)

        return round(total_limit, 3)
 # конец файла report_generator.py