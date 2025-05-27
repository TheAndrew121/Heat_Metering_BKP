import os
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill

class ReportGeneratorFrame(ttk.LabelFrame):
    def __init__(self, parent, data, cart_weights, detect_anomalies):
        super().__init__(parent, text="Формирование отчётов")
        self.parent = parent
        self.data = data
        self.cart_weights = cart_weights
        self.detect_anomalies = detect_anomalies

        # Поля выбора периода
        self.create_widgets()

    def create_widgets(self):
        # Выбор периода
        ttk.Label(self, text="Начальная дата:").grid(row=0, column=0, padx=5, pady=5)
        self.start_date_var = tk.StringVar()
        ttk.Entry(self, textvariable=self.start_date_var).grid(row=0, column=1, padx=5, pady=5)

        ttk.Label(self, text="Конечная дата:").grid(row=1, column=0, padx=5, pady=5)
        self.end_date_var = tk.StringVar()
        ttk.Entry(self, textvariable=self.end_date_var).grid(row=1, column=1, padx=5, pady=5)

        # Кнопка формирования отчёта
        ttk.Button(self, text="Сформировать отчёт", command=self.generate_report).grid(row=2, column=0, columnspan=2, pady=10)

    import os

    def generate_report(self):
        start_date = self.start_date_var.get()
        end_date = self.end_date_var.get()

        if not start_date or not end_date:
            messagebox.showerror("Ошибка", "Пожалуйста, укажите начальную и конечную дату.")
            return

        try:
            start_date_dt = pd.to_datetime(start_date, format="%d-%m-%Y")
            end_date_dt = pd.to_datetime(end_date, format="%d-%m-%Y")
        except ValueError:
            messagebox.showerror("Ошибка", "Неверный формат даты. Используйте ДД-ММ-ГГГГ.")
            return

        # Фильтрация данных
        filtered_data = [
            row for row in self.data
            if start_date_dt <= pd.to_datetime(row["date"], format="%d-%m-%Y") <= end_date_dt
        ]

        if not filtered_data:
            messagebox.showinfo("Информация", "Нет данных за указанный период.")
            return

        # Обнаружение аномалий
        anomalies = self.detect_anomalies(filtered_data)

        # Создание Excel-файла
        wb = Workbook()
        ws = wb.active
        ws.title = "Отчёт"

        # Заголовки
        headers = ["Дата", "Котельная", "Расход электроэнергии", "Общий расход воды", "Расход воды на подпитку",
                   "Тележки с углём", "Тележки с золой", "Температура подачи воды", "Температура обратной подачи воды",
                   "Температура наружного воздуха"]
        ws.append(headers)

        # Словарь для сопоставления английских ключей с русскими названиями столбцов
        key_to_header = {
            "date": "Дата",
            "boiler": "Котельная",
            "electricity": "Расход электроэнергии",
            "total_water": "Общий расход воды",
            "feed_water": "Расход воды на подпитку",
            "coal_carts": "Тележки с углём",
            "ash_carts": "Тележки с золой",
            "supply_temp": "Температура подачи воды",
            "return_temp": "Температура обратной подачи воды",
            "outdoor_temp": "Температура наружного воздуха"
        }

        # Заполнение данных
        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        for row_idx, row in enumerate(filtered_data, start=2):
            ws.append([
                row["date"], row["boiler"], row["electricity"], row["total_water"], row["feed_water"],
                row["coal_carts"], row["ash_carts"], row["supply_temp"], row["return_temp"], row["outdoor_temp"]
            ])
            for anomaly in anomalies:
                for anomaly_row_idx, anomaly_keys in anomaly.items():
                    if row_idx == anomaly_row_idx + 2:  # Смещение из-за заголовков
                        for key in anomaly_keys:
                            header = key_to_header.get(key)
                            if header:
                                col_idx = headers.index(header) + 1  # Найти индекс столбца
                                ws.cell(row=row_idx, column=col_idx).fill = red_fill

        # Автоматическая настройка ширины столбцов
        for col in range(1, len(headers) + 1):
            max_length = max(len(str(ws.cell(row=row, column=col).value)) for row in range(1, ws.max_row + 1))
            ws.column_dimensions[chr(64 + col)].width = max_length + 2  # Подстройка ширины

        # Сохранение файла
        boiler_name = filtered_data[0]["boiler"]
        file_name = f"{boiler_name} | {start_date} - {end_date}.xlsx"
        reports_dir = "reports"
        os.makedirs(reports_dir, exist_ok=True)  # Создание папки, если её нет
        file_path = os.path.join(reports_dir, file_name)
        wb.save(file_path)
        messagebox.showinfo("Успешно!", f"Отчёт сохранён в файл: {file_path}")